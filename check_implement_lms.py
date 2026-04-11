from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver.chrome.service import Service
import time
import requests
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json
import re
import pandas as pd
import tqdm
import os



#------------------------------
# Tạo file báo cáo
#------------------------------
def create_file_report(data, from_day, to_day, semester):

      wb = Workbook()
      if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
      sheet_general = wb.create_sheet("Tổng quan")
      sheet_detail = wb.create_sheet("Chi tiết")

      # TODO: Định dạng cho các dòng trong file
      header_font = Font(name="Times New Roman", size=12, bold=True)
      footer_font = Font(name="Times New Roman", size=12, bold=True)
      header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
      header_fill = PatternFill(start_color="97FFFF", end_color="97FFFF", fill_type="solid")
      data_font = Font(name="Times New Roman", size=11)
      data_alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
      border_style = Border(
            left = Side(style="thin"),
            right = Side(style="thin"),
            top = Side(style="thin"),
            bottom = Side(style="thin")
      )

      # TODO: Tạo sheet thứ 1 về tổng quản tình hình của từng khoa 
      list_department = []
      list_sum_department = []
      list_has_lms = []
      list_none_lms = []
      # TODO: tạo 3 mảng gồm khoa, tổng số nhóm, có thực hiện LMS, không có LMS
      for row in data:
            if row["department"] not in list_department:
                  list_department.append(row["department"])
                  list_sum_department.append(1)
                  list_has_lms.append(1 if row["has_lms"] != "" else 0)
                  list_none_lms.append(0 if row["has_lms"] != "" else 1)
            else:
                  idx_department = list_department.index(row["department"])
                  list_sum_department[int(idx_department)] += 1
                  if row["has_lms"] != "":
                        list_has_lms[idx_department] += 1
                  else:
                        list_none_lms[idx_department] += 1

      # TODO: Thêm dòng tiêu đề cho sheet tổng quan
      header_general_department = sheet_general.cell(row=1, column=1)
      header_general_department.value = "Khoa"
      header_general_department.font = header_font
      header_general_department.fill = header_fill
      header_general_department.alignment = header_alignment
      header_general_department.border = border_style

      header_general_sum_subject = sheet_general.cell(row=1, column=2)
      header_general_sum_subject.value = "Tổng số nhóm môn học"
      header_general_sum_subject.font = header_font
      header_general_sum_subject.fill = header_fill
      header_general_sum_subject.alignment = header_alignment
      header_general_sum_subject.border = border_style                                  
                                                      
      header_general_has_lms = sheet_general.cell(row=1, column=3)
      header_general_has_lms.value = "Đã soạn LMS"
      header_general_has_lms.font = header_font
      header_general_has_lms.fill = header_fill
      header_general_has_lms.alignment = header_alignment
      header_general_has_lms.border = border_style           

      header_general_none_lms = sheet_general.cell(row=1, column=4)
      header_general_none_lms.value = "Chưa soạn"
      header_general_none_lms.font = header_font
      header_general_none_lms.fill = header_fill
      header_general_none_lms.alignment = header_alignment
      header_general_none_lms.border = border_style    

      # TODO: Thêm dữ liệu thống kê của từng khoa
      sum_department = 0
      sum_has_lms = 0
      sum_none_lms = 0
      for depart_idx, depart in enumerate(list_department, start=2):
            body_general_department = sheet_general.cell(row=depart_idx, column=1)
            body_general_department.value = depart
            body_general_department.font = data_font
            body_general_department.alignment = data_alignment
            body_general_department.border = border_style

            # TODO depart_idx - 2: vì phải bắt đầu đổ dữ liệu ở dòng số 2, nhưng chỉ mục của list cần lấy là 0, nên cần trừ đi 2 để lấy đúng chỉ mục

            body_general_sum_subject =  sheet_general.cell(row=depart_idx, column=2)
            body_general_sum_subject.value = list_sum_department[depart_idx - 2]
            body_general_sum_subject.font = data_font
            body_general_sum_subject.alignment = data_alignment
            body_general_sum_subject.border = border_style
            sum_department += int(list_sum_department[depart_idx - 2]) # lấy tổng số nhóm môn học của tất cả các khoa

            body_general_has_lms =  sheet_general.cell(row=depart_idx, column=3)
            body_general_has_lms.value = list_has_lms[depart_idx - 2]
            body_general_has_lms.font = data_font
            body_general_has_lms.alignment = data_alignment
            body_general_has_lms.border = border_style
            sum_has_lms += int(list_has_lms[depart_idx - 2]) # lấy tổng số môn học có LMS

            body_general_none_lms = sheet_general.cell(row=depart_idx, column=4)
            body_general_none_lms.value = list_none_lms[depart_idx - 2]
            body_general_none_lms.font = data_font
            body_general_none_lms.alignment = data_alignment
            body_general_none_lms.border = border_style
            sum_none_lms += int(list_none_lms[depart_idx - 2]) # lấy tổng số môn học không có lms
      
      # TODO: thêm dòng tổng kết ở cuối sheet tổng quan
      end_header_general_department = sheet_general.cell(row=len(list_department) + 2, column=1)
      end_header_general_department.value = "Tổng cộng"
      end_header_general_department.font = footer_font
      end_header_general_department.alignment = data_alignment
      end_header_general_department.border = border_style

      end_header_general_sum_department = sheet_general.cell(row=len(list_department) + 2, column=2)
      end_header_general_sum_department.value = sum_department
      end_header_general_sum_department.font = footer_font
      end_header_general_sum_department.alignment = data_alignment
      end_header_general_sum_department.border = border_style

      end_header_general_has_lms = sheet_general.cell(row=len(list_department) + 2, column=3)
      end_header_general_has_lms.value = sum_has_lms
      end_header_general_has_lms.font = footer_font
      end_header_general_has_lms.alignment = data_alignment
      end_header_general_has_lms.border = border_style

      end_header_general_none_lms = sheet_general.cell(row=len(list_department) + 2, column=4)
      end_header_general_none_lms.value = sum_none_lms
      end_header_general_none_lms.font = footer_font
      end_header_general_none_lms.alignment = data_alignment
      end_header_general_none_lms.border = border_style
      
      # TODO: thiếp lập cột khoa về phía bên trái và điều chỉnh độ rộng của cột khoa
      max_length_header = 0
      column_department = get_column_letter(1)
      for cell in sheet_general[column_department][1:]:
            cell.alignment = Alignment(horizontal="general", vertical="center", wrap_text=True)
      for cell in sheet_general[column_department]:
            if len(str(cell.value)) > max_length_header:
                  max_length_header = len(str(cell.value))
            adjust_width_department_general = max_length_header + 5
            sheet_general.column_dimensions[column_department].width = adjust_width_department_general


      # TODO: Tạo sheet thứ 2 về chi tiết từng môn học của từng khoa
      title = [
            "STT",
            "Khoa phụ trách",
            "Mã địa phương",
            "Tên địa phương",
            "Mã môn học",
            "Tên môn học",
            "Mã nhóm",
            "Mã lớp",
            "Tên lớp",
            "Mã giảng viên",
            "Tên giảng viên",
            "Ngày bắt đầu",
            "Đã soạn LMS"
      ]

      # TODO: thiết lập giá trị và định dạng cho dòng tiêu đề
      for title_idx, row_title in enumerate(title, start = 1):
            header_detail = sheet_detail.cell(row = 1, column = title_idx)
            header_detail.value = row_title
            header_detail.font = header_font
            header_detail.alignment = header_alignment
            header_detail.fill = header_fill
            header_detail.border = border_style
      
      # TODO: thiết lập giá trị và định dạng cho các dòng còn lại
      for row_idx, row_data in enumerate(data, start = 2):
            has_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid") # mặc định các dòng trong body sẽ là màu trắng
            if row_data["has_lms"] != "x":
                  has_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") # nếu có môn chưa có lms thì sẽ được tô vàng
            
            body_detail_id = sheet_detail.cell(row = row_idx, column = 1)
            body_detail_id.value = row_idx - 1
            body_detail_id.font = data_font
            body_detail_id.alignment = data_alignment
            body_detail_id.border = border_style
            body_detail_id.fill = has_fill

            body_detail_department = sheet_detail.cell(row = row_idx, column = 2)
            body_detail_department.value = row_data["department"]
            body_detail_department.font = data_font
            body_detail_department.alignment = data_alignment
            body_detail_department.border = border_style
            body_detail_department.fill = has_fill

            body_detail_id_unit = sheet_detail.cell(row = row_idx, column = 3)
            body_detail_id_unit.value = row_data["id_unit"]
            body_detail_id_unit.font = data_font
            body_detail_id_unit.alignment = data_alignment
            body_detail_id_unit.border = border_style
            body_detail_id_unit.fill = has_fill

            body_detail_name_unit = sheet_detail.cell(row = row_idx, column = 4)
            body_detail_name_unit.value = row_data["name_unit"]
            body_detail_name_unit.font = data_font
            body_detail_name_unit.alignment = data_alignment
            body_detail_name_unit.border = border_style
            body_detail_name_unit.fill = has_fill

            body_detail_id_subject = sheet_detail.cell(row = row_idx, column = 5)
            body_detail_id_subject.value = row_data["id_subject"]
            body_detail_id_subject.font = data_font
            body_detail_id_subject.alignment = data_alignment
            body_detail_id_subject.border = border_style
            body_detail_id_subject.fill = has_fill

            body_detail_name_subject = sheet_detail.cell(row = row_idx, column = 6)
            body_detail_name_subject.value = row_data["name_subject"]
            body_detail_name_subject.font = data_font
            body_detail_name_subject.alignment = data_alignment
            body_detail_name_subject.border = border_style
            body_detail_name_subject.fill = has_fill

            body_detail_group = sheet_detail.cell(row = row_idx, column = 7)
            body_detail_group.value = row_data["group"]
            body_detail_group.font = data_font
            body_detail_group.alignment = data_alignment
            body_detail_group.border = border_style
            body_detail_group.fill = has_fill

            body_detail_id_class = sheet_detail.cell(row = row_idx, column = 8)
            body_detail_id_class.value = row_data["id_class"]
            body_detail_id_class.font = data_font
            body_detail_id_class.alignment = data_alignment
            body_detail_id_class.border = border_style
            body_detail_id_class.fill = has_fill

            body_detail_name_class = sheet_detail.cell(row = row_idx, column = 9)
            body_detail_name_class.value = row_data["name_class"]
            body_detail_name_class.font = data_font
            body_detail_name_class.alignment = data_alignment
            body_detail_name_class.border = border_style
            body_detail_name_class.fill = has_fill

            body_detail_id_teacher = sheet_detail.cell(row = row_idx, column = 10)
            body_detail_id_teacher.value = row_data["id_teacher"]
            body_detail_id_teacher.font = data_font
            body_detail_id_teacher.alignment = data_alignment
            body_detail_id_teacher.border = border_style
            body_detail_id_teacher.fill = has_fill

            body_detail_name_teacher = sheet_detail.cell(row = row_idx, column = 11)
            body_detail_name_teacher.value = row_data["name_teacher"]
            body_detail_name_teacher.font = data_font
            body_detail_name_teacher.alignment = data_alignment
            body_detail_name_teacher.border = border_style
            body_detail_name_teacher.fill = has_fill

            body_detail_from_day = sheet_detail.cell(row = row_idx, column = 12)
            body_detail_from_day.value = row_data["from_day"]
            body_detail_from_day.font = data_font
            body_detail_from_day.alignment = data_alignment
            body_detail_from_day.border = border_style
            body_detail_from_day.fill = has_fill

            body_detail_has_lms = sheet_detail.cell(row = row_idx, column = 13)
            body_detail_has_lms.value = row_data["has_lms"]
            body_detail_has_lms.font = data_font
            body_detail_has_lms.alignment = data_alignment
            body_detail_has_lms.border = border_style
            body_detail_has_lms.fill = has_fill

      # TODO: điều chỉnh độ rộng của cột dựa trên giá trị dài nhất
      for col_idx in range(1, len(data[0]) + 1):
            max_length = 0
            column = get_column_letter(col_idx)
            for cell in sheet_detail[column]:
                  if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            adjust_width = max_length + 2
            sheet_detail.column_dimensions[column].width = adjust_width
      
      root_folder = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
      save_files_folder = os.path.join(root_folder, semester)
      if os.path.exists(save_files_folder):
            os.makedirs(save_files_folder)

      #name_file = f"{semester} - Tình hình soạn thảo LMS từ ngày {datetime.strptime(from_day, "%Y-%m-%d").strftime("%d-%m-%Y")} đến ngày {datetime.strptime(to_day, "%Y-%m-%d").strftime("%d-%m-%Y")}.xlsx"
      name_file = f"BC Tình hình soạn thảo LMS HK {semester} từ ngày ({datetime.strptime(from_day, "%Y-%m-%d").strftime("%d-%m-%Y")} đến ngày {datetime.strptime(to_day, "%Y-%m-%d").strftime("%d-%m-%Y")}).xlsx"
      wb.save(name_file)


