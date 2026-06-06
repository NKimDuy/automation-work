from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver.chrome.service import Service
from modules.init_selenium import InitSelenium
from utils.config_loader import ConfigLoader
from utils.logger import setup_logger
# from db.connect_mysql import get_connection
from config.settings import semester
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
import os
import time
import re
import copy
import joblib
import re


class ScoreLMS:
    def __init__(self):
        # self.url_lms = f"https://lms.oude.edu.vn/{semester}/course/search.php"
        self.base_dir = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
        self.url_lms = f"https://lms.oude.edu.vn/251/course/search.php"
        self.DIC_SCORE_BASE = {
            "thlms": {"score": "", "name": "Thực hiện LMS"},
            "ctdd": {"score": 0, "name": "Cấu trúc đầy đủ"},
            "ttkh": {"score": 0, "name": "Thông tin về khóa học"},
            "mtht": {"score": 0, "name": "Mục tiêu, hoạt động học tập"},
            "ttbg": {"score": 0, "name": "Thông tin bài giảng"},
            "bdgk": {"score": 0, "name": "Bảng điểm giữa kỳ"},
            "cbg": {"score": "", "name": "Video ghi hình buổi giảng"},
            "tltk": {"score": "", "name": "Tài liệu tham khảo/bài đọc thêm"},
            "bt": {"score": 0, "name": "Bài tập tự luận/trắc nghiệm"},
            "dd": {"score": "", "name": "Diễn đàn"},
            "vgdtm": {"score": "", "name": "video giải đáp thắc mắc cho sinh viên"},
            "svdk": {"score": 0, "name": "Sinh viên đăng ký"},
            "svtc": {"score": 0, "name": "Sinh viên truy cập"},
            "tlsvtc": {"score": 0, "name": "Tỷ lệ sinh viên truy cập"},
            "tlgvtc": {"score": 0, "name": "Tỷ lệ giảng viên truy cập"},
            "td": {"score": 0, "name": "Tổng điểm"},
            "kqdg": {"score": "", "name": "kết quả đánh giá"},
        }
        self.DIC_CRETERIA = {
            1: "Giới thiệu môn học",
            2: "Giới thiệu giảng viên",
            3: "Đề cương môn học",
            4: "Phương thức kiểm tra, đánh giá môn học",
            5: "Tỷ lệ điểm đánh giá",
            6: "Giới thiệu ngắn về mục tiêu, hoạt động học tập",
            7: "Nội dung bài giảng",
            8: "Tài liệu tham khảo/bài đọc thêm",
            9: "Video giảng online",
            10: "Bảng điểm giữa kỳ",
        }
        self.NOT_RATED = [
            "Thí nghiệm vật liệu xây dựng",
            "Thí nghiệm cơ chất lỏng",
            "Thí nghiệm sức bền vật liệu",
            "Thực tập trắc địa",
            "Địa chất công trình + thực tập",
            "Thí nghiệm cơ học đất",
            "Nhận thức ngành",
            'Thực tập địa chất công trình',
            "Đồ án nền móng",
            "Đồ án kết cấu bê tông cốt thép 1, 2",
            "Đồ án bê tông 1",
            "Đồ án bê tông 2",
            "Đồ án thi công",
            "Đồ án kết cấu thép",
            "Đồ án cấp thoát nước trong nhà",
            "Đồ án cấp thoát nước",
            "Đồ án mạng lưới cấp thoát nước",
            "Đồ án tổ chức và quản lý thi công",
            "Đồ án quản lý dự án xây dựng",
            "Đồ án lập và thảm định dự án ĐTXD",
            "Đồ án phân tích định lượng trong QLXD",
            "Thực tập tốt nghiệp",
            "Đồ án tốt nghiệp"
        ] 
        self.NOISE_REMOVE = [
            'rút gọn',
            "các thông báo",
            "bài tập",
            "câu hỏi thảo luận",
            "diễn đàn",
            'Chuyển tới nội dung chính',
            'LMS ĐTTX',
            'Topic outline',
            'Thu gọn toàn bộ',
            'Copyright © 2024. Powered by Moodle'
        ]

    def preprocess(self, text: str) -> str:
        if not isinstance(text, str):
            return ""
        text = text.lower().strip()
        # Xoá ký tự đặc biệt, giữ chữ cái và khoảng trắng
        text = re.sub(r"[^\w\sàáâãèéêìíòóôõùúýăđơưạảấầẩẫậắằẳẵặẹẻẽếềểễệỉịọỏốồổỗộớờởỡợụủứừửữựỳỵỷỹ]", " ", text)
        text = re.sub(r"\s+", " ", text).strip()
        return text
    
    def predict(self, model, texts: list) -> list:
        cleaned = [self.preprocess(t) for t in texts]
        preds = model.predict(cleaned)
        results = list()
        for label in preds:
            results.append(label)
        return results
    
    def check_forum(self, driver, forum_link, ten_giang_vien):
        second_tab = driver.current_window_handle # lưu tab môn học, vì khi mở diễn đàn, sẽ mở sang tab khác
        driver.execute_script("window.open(arguments[0]);", forum_link)
        driver.switch_to.window(driver.window_handles[-1])

        rows = driver.find_elements(By.CSS_SELECTOR, "tr.discussion")
        flad_forum = False
        if rows:
            for row in rows:
                try:
                    number_responses = row.find_element(By.CSS_SELECTOR, "td.text-center span").text.strip()
                    if int(number_responses) == 1:
                        # try:
                        #     ten_sv = row.find_element(By.CSS_SELECTOR, "td.author .mb-1").text.strip()
                        # except:
                        #     ten_sv = "Không rõ"
                        try:
                            ten_reply = row.find_element(By.CSS_SELECTOR, "td.text-start .mb-1").text.strip()
                        except:
                            ten_reply = "Không có"

                        if ten_giang_vien.lower() not in ten_reply.lower():
                            flad_forum = True
                            break
                    else:
                        try:

                            thrid_tab = driver.current_window_handle

                            link_topic = row.find_element(By.CSS_SELECTOR, "th.topic a").text.strip()
                            link_topic = link_topic.get_attribute("href")
                            driver.execute_script("window.open(arguments[0]);", link_topic)
                            driver.switch_to.window(driver.window_handles[-1])
                            try:
                                list_teacher_name = []
                                find_teacher_name = driver.find_elements(By.CSS_SELECTOR, "header .mb-2")
                                for item in find_teacher_name:
                                    teacher_name = item.find_element(By.CSS_SELECTOR, "div.mb-3 a").text.strip()
                                    list_teacher_name.append(teacher_name.lower())
                                driver.switch_to.window(thrid_tab)
                                if list_teacher_name.count(ten_giang_vien.lower()) > 1:
                                    flad_forum = True
                                    break 
                            except:
                                ten_gv = "Không rõ"

                        except:
                            print("Không tìm thấy link topic trong diễn đàn")

                except:
                    number_responses = 0

            
        driver.close()
        driver.switch_to.window(second_tab)
        return flad_forum

    def get_results_score(self, driver, model, ten_gv, list_lsa):
        notes = '' # lưu các tiêu chí chưa đạt để ghi chú vào file excel
        sum_score = 0 # lưu tổng điểm của môn học, sẽ được ghi vào file excel ở cột "Tổng điểm"
        # copy từ điển điểm để áp dụng cho từng môn học, tránh bị ghi đè khi duyệt qua nhiều môn
        dic_score_apply = copy.deepcopy(self.DIC_SCORE_BASE)

        forums = {} # chứa thông tin diễn đàn: {ten_forum: link_forum}
        quiz = {}  # chứa thông tin bài tập quiz: {ten_quiz: link_quiz}
        assign = {} # chứa thông tin bài tập tự luận: {ten_assign: link_assign}

        dic_score_apply["thlms"]["score"] = "x"
        dic_score_apply["ctdd"]["score"] = 15
        sum_score += 15

        # lấy tất cả text có trong môn học 
        full_text = driver.find_element(By.TAG_NAME, "body").text
        # tách text thành từng dòng để model dự đoán,
        #vì model được train trên từng câu, nên cần tách text thành từng dòng/câu để dự đoán chính xác hơn
        lines = full_text.splitlines()  
        # loại bỏ các dòng có 1 từ hoặc không có từ nào, vì những dòng này thường không mang nhiều thông tin
        lines = [item for item in lines if len(item.split()) > 1] 
        # loại bỏ các dòng có chứa các từ khóa không liên quan đến việc đánh giá cấu trúc môn học, 
        # như "bài tập", "diễn đàn", "câu hỏi thảo luận", 
        # vì những dòng này thường là các phần tử phụ trong LMS và không phải là tiêu chí đánh giá cấu trúc môn học
        lines = [line for line in lines if not any(noise in line.lower() for noise in self.NOISE_REMOVE)]
        results = self.predict(model, lines)
        results = list(set(results))  # lấy kết quả duy nhất để đánh giá

        get_ttkh = [1, 2, 3, 4, 5] 
        check_ttkn = [x for x in get_ttkh if x not in set(results)]
        if not check_ttkn:
            dic_score_apply["ttkh"]["score"] = 10
            sum_score += 10
        else:
            # notes += "\nChưa đủ thông tin khóa học, cụ thể ("
            name_not_check_ttkn = ["thiếu " + self.DIC_CRETERIA[i] for i in check_ttkn]
            # notes += " ".join(name_not_check_ttkn)
            # # notes += "\n".join(name_not_check_ttkn)
            # notes += ")"
            notes += f"\nChưa đủ thông tin khóa học, cụ thể ({','.join(name_not_check_ttkn)})"
            
        
        if 6 in results:
            dic_score_apply["mtht"]["score"] = 20
            sum_score += 20
        else:
            notes += "\nThiếu giới thiệu về mục tiêu học tập"
        if 7 in results:
            dic_score_apply["ttbg"]["score"] = 20
            sum_score += 20
        else:
            notes += "\nThiếu nội dung bài giảng"
        if 8 in results:
            dic_score_apply["tltk"]["score"] = "x"
        else:
            notes += "\nThiếu tài liệu tham khảo/bài đọc thêm"
        if 9 in results:
            dic_score_apply["cbg"]["score"] = "x"
        if 10 in results:
            dic_score_apply["bdgk"]["score"] = 15
            sum_score += 15
        else:
            notes += "\nThiếu bảng điểm giữa kỳ"

        # khúc này sẽ gọi lsa để lấy dữ liệu sinh viên đăng ký và sinh viên truy cập,
        #  từ đó tính ra tỷ lệ sinh viên truy cập và tỷ lệ giảng viên truy cập, 
        # nếu có dữ liệu thì sẽ áp dụng điểm số tương ứng, nếu không có dữ liệu thì sẽ ghi chú vào file excel
        dic_score_apply["svdk"]["score"] = list_lsa[0] if list_lsa else "lỗi"
        dic_score_apply["svtc"]["score"] = list_lsa[1] if list_lsa else "lỗi"
        dic_score_apply["tlsvtc"]["score"] = list_lsa[2] if list_lsa else "lỗi"
        dic_score_apply["tlgvtc"]["score"] = list_lsa[3] if list_lsa else "lỗi"

        all_links = driver.find_elements(By.TAG_NAME, "a")
        for a in all_links:
            a_href = a.get_attribute("href")
            a_text = a.text.strip()
            
            if "forum" in a_href:
                forums[a_text] = a_href
            if "quiz" in a_href:
                quiz[a_text] = a_href
            if "assign" in a_href:
                assign[a_text] = a_href

        check_bt = [name for name, d in [("quiz", quiz), ("assign", assign)] if d]
        if check_bt:
            dic_score_apply["bt"]["score"] = 20
            sum_score += 20
        else:
            notes += "\nThiếu bài tập"

        # check_bt = [name for name, d in [("quiz", quiz), ("assign", assign)] if not d]
        # if not check_bt:
        #     dic_score_apply["bt"]["score"] = 20
        #     sum_score += 20
        # else:
        #     if check_bt == ["quiz"]:
        #         notes += "\nThiếu bài tập trắc nghiệm"
        #     elif check_bt == ["assign"]:
        #         notes += "\nThiếu bài tập tự luận"
        #     else:
        #         notes += "\nThiếu bài tập tự luận/trắc nghiệm"

        dic_score_apply["td"]["score"] = sum_score

        # đánh giá diễn đàn, nếu có diễn đàn thì sẽ mở từng diễn đàn lên để kiểm tra xem giảng viên đã tương tác với tất cả sinh viên chưa, nếu có diễn đàn nhưng không tương tác đầy đủ thì sẽ ghi chú vào file excel, nếu có diễn đàn và tương tác đầy đủ thì sẽ được điểm tối đa, nếu không có diễn đàn nào thì sẽ không được điểm nào và ghi chú vào file excel
        if forums:
            forum_chua_day_du = []  # lưu các diễn đàn chưa tương tác đầy đủ [(ten_forum, [sv_chua_reply])]
            for name, forum_link in forums.items():
                chua_reply = self.check_forum(driver, forum_link, ten_gv)
                if chua_reply:
                    forum_chua_day_du.append({
                        "ten_forum": name
                    })
            if not forum_chua_day_du:
                #print(f" ✅ Có diễn đàn và tương tác với tất cả sinh viên")
                dic_score_apply["dd"]["score"] = "x"
            else:
               # print(f" ⚠️ Có diễn đàn nhưng không tương tác đầy đủ với sinh viên")
                for f in forum_chua_day_du:
                    #print(f"  📌 Forum: {f['ten_forum']}")
                    notes += f"\nDiễn đàn '{f['ten_forum']}' chưa tương tác đầy đủ với sinh viên"
        else:
            #print(f" ❌ Không có diễn đàn nào")
            notes += "\nKhông có diễn đàn"   

        score_basic = [
            dic_score_apply["ctdd"]["score"] == 15,
            dic_score_apply["ttkh"]["score"] == 10,
            dic_score_apply["mtht"]["score"] == 20,
            dic_score_apply["ttbg"]["score"] == 20,
            dic_score_apply["bdgk"]["score"] == 15,
            float(dic_score_apply["tlsvtc"]["score"].replace('%', '')) >= 70,
            float(dic_score_apply["tlgvtc"]["score"].replace('%', '')) > 0,
            sum_score >= 50
        ]

        if all(score_basic):
            dic_score_apply["kqdg"]["score"] = "Đạt cơ bản"

            score_advance = [
                dic_score_apply["dd"]["score"] == "x",
                dic_score_apply["bt"]["score"] == 20,
                float(dic_score_apply["tlsvtc"]["score"].replace('%', '')) >= 90,
                float(dic_score_apply["tlgvtc"]["score"].replace('%', '')) > 0,
                sum_score >= 80
            ]
            if all(score_advance):
                dic_score_apply["kqdg"]["score"] = "Xem xét nâng cao"

        else:
            dic_score_apply["kqdg"]["score"] = "Chưa đạt cơ bản"

        return [dic_score_apply, notes] 

    def test_lsa(self):

        dic_lsa = {}
        start_selenium = InitSelenium()
        get_lsa_data = start_selenium.process_get_detail_lsa()
        for d in get_lsa_data:
            if "subject_id" not in d:
                continue
            key = f"{d['subject_id']}_{d['group_id']}_{d['teacher_id']}"
            value = [d["count_total_student"], d["count_student_access"], d["percent_student_access"], d["percent_teacher_access"]]
            dic_lsa[key] = value  
        print(dic_lsa)  
        
        excel_file = os.path.join(self.base_dir, "data", "input", "score_lms","251","test.xlsx")
        wb_score = load_workbook(excel_file)
        ws_score = wb_score.active
        for i, row in enumerate(ws_score.iter_rows(min_row=2, values_only=True), start=2):
            ten_subject = row[0]  # Giả sử tên môn học nằm ở cột A
            ten_gv = row[1]       # Giả sử tên giảng viên nằm ở cột B 
            ma_mh = row[2]       # Giả sử mã môn học nằm ở cột C
            ma_nhom = row[3]      # Giả sử mã nhóm môn học nằm ở cột D
            ma_gv = row[4]       # Giả sử mã giảng viên nằm ở cột E
            
            key_lsa = f"{ma_mh}_{ma_nhom}_{ma_gv}"

            print(dic_lsa.get(key_lsa))


    def score_lms(self):
        
        model_file = os.path.join(self.base_dir, "AI", "model", "model_vi_classification_v2.pkl")
        model = joblib.load(model_file)

        excel_file = os.path.join(self.base_dir, "data", "input", "score_lms","251","test.xlsx")
        # excel_file = os.path.join(self.base_dir, "data", "input", "score_lms","251","2025-06-04_251_email-nhom-mh-hk.xlsx")
        wb_score = load_workbook(excel_file)
        ws_score = wb_score.active

        # lần đầu kiểm tra trong file đã có các cột để tiến hành đánh giá chưa
        # Lấy cột thực hiện lms kiểm tra (lấy bất kỳ cột nào cũng được) nhưng lấy cột thực hiện lms
        # vì cột này sẽ luôn đứng sau các cột có sẵn trong file
        headers = [ws_score.cell(row=1, column=col).value for col in range(1, ws_score.max_column + 1)]
        column_current = ws_score.max_column               
        if self.DIC_SCORE_BASE["thlms"]["name"] not in headers:
            column_add = column_current + 1
            for col_criteria in self.DIC_SCORE_BASE:
                ws_score.cell(row=1, column=column_add).value = self.DIC_SCORE_BASE[col_criteria]["name"]
                column_add += 1
            ws_score.cell(row=1, column=column_add).value = "Ghi chú"

        start_selenium = InitSelenium()
        driver = start_selenium.login_selenium(self.url_lms)

        # tạo dictionary với key là mã môn học + mã nhóm + mã giảng viên, value là list chứa dữ liệu lsa
        # dic_lsa = {}
        # get_lsa_data = start_selenium.process_get_detail_lsa()
        # for d in get_lsa_data:
        #     if "subject_id" not in d:
        #         continue
        #     key = f"{d['subject_id']}_{d['group_id']}_{d['teacher_id']}"
        #     value = [d["count_total_student"], d["count_student_access"], d["percent_student_access"], d["percent_teacher_access"]]
        #     dic_lsa[key] = value
        temp = [2,2,'20%', '20%']
        #chỗ này sẽ chạy vòng lặp for để duyệt qua các môn học trong file excel
        for i, row in enumerate(ws_score.iter_rows(min_row=2, values_only=True), start=2):
            if ws_score.cell(row=i, column=column_current + len(self.DIC_SCORE_BASE) + 2).value != "x":
                # mỗi lần duyệt qua một môn học thì sẽ mở lại trang tìm kiếm môn học trên LMS 
                # để đảm bảo rằng các bước tìm kiếm và đánh giá được thực hiện chính xác cho từng môn học, 
                # tránh bị lỗi do trang web bị thay đổi khi duyệt qua nhiều môn học
                driver.get(self.url_lms)

                # subject_id = row[0] 
                # subject_name = row[1]  
                # teacher_id = row[8]
                # teacher_name = row[9]      
                # group_id = row[6] 
                #
                subject_id = row[2] 
                subject_name = row[0]  
                teacher_id = row[4]
                teacher_name = row[1]      
                group_id = row[3]      
                    
                # key_lsa = f"{subject_id}_{group_id}_{teacher_id}"
                serch_subject = f"{subject_id} - ({teacher_id}-{group_id})"

                

                # kiểm tra xem môn học có thuộc danh sách các môn không đánh giá điểm LMS của khoa xây dựng hay không, 
                # nếu có thì sẽ ghi "Hoàn thành" vào cột điểm số và bỏ qua các bước kiểm tra tiếp theo, 
                # vì những môn này chỉ cần soạn LMS hoàn chỉnh là được, không yêu cầu phải có đầy đủ các tiêu chí như các môn khác
                if subject_name.lower() in [subject.lower() for subject in self.NOT_RATED]:
                    # print(f"✅ Môn học thuộc cách chấm chỉ hoàn thành của khoa xây dựng")
                    ws_score.cell(row=i, column=column_current + len(self.DIC_SCORE_BASE)).value = "Hoàn thành"

                    print(f"kết quả xử lý môn {subject_name} - {teacher_name} - {group_id}: Hoàn thành")
                    continue

                try:
                    search_input = WebDriverWait(driver, 10).until(
                        EC.visibility_of_element_located((By.NAME, "q"))
                    )
                    search_input.clear()
                    search_input.send_keys(serch_subject) # BADM1391 - (QT653-TN120)
                except Exception as e:
                        print(f"Không tìm thấy ô nhập môn học, lỗi {e}")

                try:
                    # Nhấn nút tìm kiếm
                    button_find_text = "button[type='submit']"
                    button_find = WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, button_find_text))
                    )
                    button_find.click()
                except Exception as e:
                        print(f"Không tìm thấy nút tìm môn học, lỗi {e}")

                try: 
                    # Lấy tất cả link môn học trong kết quả tìm kiếm (class="aalink")
                    get_list_subject = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.CLASS_NAME, "aalink"))
                    )
                    href_link_subject = get_list_subject.get_attribute("href")
                    # mở link môn học để kiểm tra các tiêu chí đánh giá cấu trúc môn học trên LMS
                    driver.get(href_link_subject)
                    # gọi hàm get_results_score để đánh giá các tiêu chí và lấy điểm số áp dụng cho từng tiêu chí, 
                    # cũng như ghi chú các tiêu chí chưa đạt vào file excel
                    # dic_score_apply, notes = self.get_results_score(driver, model, teacher_name, dic_lsa.get(key_lsa))
                    dic_score_apply, notes = self.get_results_score(driver, model, teacher_name, temp)

                    print(f"kết quả xử lý môn {subject_name} - {teacher_name} - {group_id}: {dic_score_apply['kqdg']['score']}")
                    
                    col_idx = column_current + 1
                    # dic_score_apply sẽ trả về dictionary
                    # for col_criteria in dic_score_apply: sẽ duyệt qua các key của dictionary
                    for col_criteria in dic_score_apply:
                        ws_score.cell(row=i, column=col_idx).value = dic_score_apply[col_criteria]["score"]
                        col_idx += 1
                    ws_score.cell(row=i, column=col_idx).value = notes
                    ws_score.cell(row=i, column=col_idx).alignment = Alignment(wrap_text=True)             
                
                    driver.get(self.url_lms) # quay lại trang tìm kiếm để tiếp tục duyệt qua môn học tiếp theo
                    time.sleep(2)  # Thêm delay để dừng lại 2 giây trước khi nhập thông tin môn
                except Exception as e:
                    # nếu không tìm thấy môn học trên LMS thì sẽ ghi "Chưa soạn LMS" vào cột điểm số và ghi chú vào file excel, sau đó tiếp tục chạy vòng lặp để duyệt qua môn học tiếp theo
                    ws_score.cell(row=i, column=column_current + len(self.DIC_SCORE_BASE)).value = "Chưa soạn LMS"
                    print(f"kết quả xử lý môn {subject_name} - {teacher_name} - {group_id}: Chưa soạn LMS")
                ws_score.cell(row=i, column=column_current + len(self.DIC_SCORE_BASE) + 2).value = "x"
        wb_score.save(excel_file)     
        driver.quit()
        
test = ScoreLMS()
test.score_lms()
# print(test.test_lsa())