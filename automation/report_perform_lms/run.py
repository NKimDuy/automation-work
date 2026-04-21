# =============================================================================
# MODULE: report_perform_lms/run.py
# MỤC ĐÍCH: Tạo báo cáo tình hình soạn thảo LMS theo tuần.
#
# LUỒNG CHẠY CHÍNH:
#   1. Lấy thông tin giảng viên từ PHDT (scraping Selenium)
#   2. Lấy danh sách môn học đã có trên LSA (scraping Selenium)
#   3. Gọi API lấy toàn bộ môn học trong học kỳ theo từng đơn vị
#   4. Lọc các môn có ngày bắt đầu TKB nằm trong khoảng tuần cần báo cáo
#   5. Đối chiếu xem môn đó đã được ghi danh trên LSA chưa → đánh dấu "x"
#   6. Xuất file Excel báo cáo (sheet Tổng quan + sheet Chi tiết)
#   7. Ghi môn chưa soạn LMS vào file theo dõi tích lũy
#
# FILE ĐẦU RA:
#   - data/output/report_perform_lms/<semester>/
#       BC Tình hình soạn thảo LMS HK <semester> từ ngày (...).xlsx  → báo cáo tuần
#       __danh_sach_khong_thuc_hien_lms.xlsx                          → file tích lũy các môn chưa soạn
#
# CHẠY TRỰC TIẾP:
#   python -m automation.report_perform_lms.run
# =============================================================================

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver.chrome.service import Service
from modules.init_selenium import InitSelenium
from utils.config_loader import ConfigLoader
from utils.logger import setup_logger
from utils.api import APIHandler
from config.settings import semester
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import time
from datetime import datetime, timedelta
import re


class ReportPerformLMS:

      def test(self):
            """Hàm thử nghiệm: in ra danh sách nhóm-môn học lấy được từ LSA.

            Dùng để kiểm tra nhanh xem scraping LSA có hoạt động đúng không.
            Không dùng trong luồng chính.
            """
            start_selenium = InitSelenium()
            # Lấy danh sách môn học từ LSA thông qua Selenium
            list_lsa = start_selenium.process_get_detail_lsa()
            try:
                  for ls in list_lsa:
                        if ls["group_id"] and ls["subject_id"]:
                              # In ra theo định dạng "nhóm-mã môn", ví dụ: "SG001-ACCO4331"
                              print("-".join([ls["group_id"], ls["subject_id"]]))
            except Exception as e:
                  print(f"An error occurred: {e}")


      def get_last_week_range(self):
            """Tính khoảng ngày của tuần trước (thứ 2 → chủ nhật).

            Hàm này được dùng khi chạy báo cáo vào thứ 2 hàng tuần —
            mục đích là lấy dữ liệu của tuần vừa rồi, không phải tuần hiện tại.

            Ví dụ: Hôm nay là thứ 4 ngày 16/04/2026
                   → Trả về ("2026-04-06", "2026-04-12")  ← thứ 2 và CN tuần trước

            Trả về:
                tuple(str, str): (ngày_thứ_2_tuần_trước, ngày_chủ_nhật_tuần_trước)
                                  định dạng "YYYY-MM-DD"
            """
            today = datetime.today()

            # weekday() trả về: 0=Thứ 2, 1=Thứ 3, ..., 6=Chủ nhật
            # Lùi về thứ 2 tuần này (weekday: 0=T2, 6=CN)
            days_since_monday = today.weekday()  # hôm nay là thứ mấy tính từ T2

            last_monday = today - timedelta(days=days_since_monday + 7)  # T2 tuần trước
            last_sunday = last_monday + timedelta(days=6)                # CN tuần trước

            return last_monday.strftime("%Y-%m-%d"), last_sunday.strftime("%Y-%m-%d")


      def get_department_of_subject(self):
            """Đọc file Excel ánh xạ mã môn học → tên khoa từ file cấu hình.

            File Excel nằm tại: config/excel/mon_hoc_khoa.xlsx
            Cột 0: Mã môn học (ví dụ: "ACCO4331")
            Cột 1: Mã khoa   (ví dụ: "TX.KKKK")

            Trả về:
                dict: { mã_môn_học: mã_khoa, ... }
                      ví dụ: { "ACCO4331": "TX.KKKK", "LAWS1234": "TX.LALA" }

            Lưu ý: Hàm này được gọi lặp lại trong vòng lặp lớn ở report_perform_lms.
                   Nếu file lớn và chậm, có thể cache kết quả lại bằng cách gán vào self.
            """
            # Tìm đường dẫn tuyệt đối đến thư mục gốc của project (lên 3 cấp từ file này)
            base_dir = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
            file = os.path.join(base_dir, "config", "excel", "mon_hoc_khoa.xlsx")
            wb = load_workbook(file)
            ws = wb.active
            dict_subject_department = {}
            # Bỏ qua hàng đầu tiên (header), bắt đầu đọc từ hàng 2
            for row in ws.iter_rows(min_row=2, values_only=True):
                  subject_code = row[0]   # cột A: mã môn học
                  department = row[1]     # cột B: mã khoa
                  dict_subject_department[subject_code] = department
            return dict_subject_department


      def report_perform_lms(self, from_day, to_day):
            """Pipeline chính: tổng hợp và xuất báo cáo tình hình soạn thảo LMS.

            Tham số:
                from_day (str): ngày bắt đầu khoảng báo cáo, định dạng "YYYY-MM-DD"
                to_day   (str): ngày kết thúc khoảng báo cáo, định dạng "YYYY-MM-DD"

            Gọi thủ công (run.py) hoặc tự động (main.py):
                from_day, to_day = ob.get_last_week_range()
                ob.report_perform_lms(from_day, to_day)

            Các bước xử lý:
              1. Scraping PHDT → lấy thông tin giảng viên theo từng nhóm-môn
              2. Scraping LSA  → lấy danh sách nhóm-môn đã được ghi danh trên LMS
              3. Gọi API       → lấy toàn bộ môn học trong học kỳ từng đơn vị
              4. Lọc môn có TUNGAYTKB nằm trong khoảng from_day → to_day
              5. Tra khoa theo mã môn (từ file Excel cấu hình)
              6. Đối chiếu với danh sách LSA → đánh dấu "x" nếu đã soạn LMS
              7. Gọi decor_report_perform_lms() để xuất file Excel
            """

            # --- BƯỚC 1: Lấy thông tin giảng viên từ PHDT ---
            # process_get_detail_phdt() trả về dict dạng:
            #   { "nhóm-mã_môn": [mã_giảng_viên, tên_giảng_viên], ... }
            # Ví dụ: { "SG001-ACCO4331": ["TX001", "Nguyễn Văn A"] }
            start_selenium = InitSelenium()
            get_info_teacher = start_selenium.process_get_detail_phdt()

            # --- BƯỚC 2: Lấy danh sách môn học đã ghi danh trên LSA ---
            # process_get_detail_lsa() trả về list các dict, mỗi dict có:
            #   lms_id, subject_id, subject_name, teacher_id, teacher_name, group_id, ...
            list_lsa = start_selenium.process_get_detail_lsa()

            # Tạo list các key "nhóm-mã_môn" để đối chiếu nhanh sau này
            # Ví dụ: ["SG001-ACCO4331", "SG002-LAWS1234", ...]
            get_group_subject_in_lsa = []
            try:
                  for item in list_lsa:
                        key = "-".join([item["group_id"], item["subject_id"]])
                        get_group_subject_in_lsa.append(key)
            except Exception as e:
                  print(f"Không có dữ liệu môn học: {e}")

            # --- BƯỚC 3: Gọi API lấy toàn bộ môn học theo từng đơn vị ---
            api_handler = APIHandler()
            # get_unit() trả về [[MaDP, TenDP], ...] — danh sách các đơn vị (khoa/trung tâm)
            list_unit = api_handler.get_unit()

            # Chuyển chuỗi ngày sang datetime để so sánh
            from_day = datetime.strptime(from_day, "%Y-%m-%d")
            to_day = datetime.strptime(to_day, "%Y-%m-%d")

            # Bảng ánh xạ mã khoa (từ file Excel) sang tên khoa hiển thị trong báo cáo
            # Mã khoa dạng "TX.XXXX" là mã nội bộ, cần đổi sang tên đọc được
            department_dic = {
                  "TX.NNNN": "Ngoại ngữ",
                  "TX.LALA": "Luật",
                  "TX.LA": "Luật",
                  "TX.CBML": "Cơ bản",
                  "TX.CBCB": "Cơ bản",
                  "TX.XHXH": "XHH-CTXH-ĐNA",
                  "TX.KKKK": "Kế toán - Kiểm toán",
                  "TX.QTQT": "Quản Trị Kinh Doanh",
                  "TX.TCTC": "Tài chính - Ngân hàng",
                  "TX.SHSH": "Công Nghệ Sinh Học",
                  "TX.KIKI": "Kinh tế và quản lý công",
                  "TX.KTKT": "Xây dựng",
            }

            # --- BƯỚC 4: Lọc môn học theo khoảng ngày TKB và tổng hợp dữ liệu ---
            # list_report_lms là dict dùng key "nhóm-mã_môn" để tránh trùng lặp
            # Value là list: [khoa, nhóm, mã_môn, tên_môn, mã_lớp, mã_ĐV, tên_ĐV,
            #                 mã_GV, tên_GV, ngày_bắt_đầu_TKB]
            list_report_lms = {}

            for unit in list_unit:
                  # Lấy danh sách môn học theo từng đơn vị và học kỳ hiện tại
                  list_subject = api_handler.get_subject_from_api(semester, unit[0])
                  for subject in list_subject:
                        # Bỏ qua môn chưa có ngày bắt đầu TKB
                        if subject["TUNGAYTKB"] is not None:
                              # Chỉ lấy môn có TUNGAYTKB nằm trong khoảng tuần cần báo cáo
                              if from_day <= datetime.strptime(subject["TUNGAYTKB"], "%Y-%m-%d") <= to_day:
                                    # Tra mã khoa từ file Excel, sau đó đổi sang tên khoa hiển thị
                                    key_department = self.get_department_of_subject().get(subject["MaMH"], "Không xác định")
                                    value_department = department_dic.get(key_department, "Không xác định")

                                    # Key dùng để nhóm: "nhóm-mã_môn", ví dụ "SG001-ACCO4331"
                                    key = "-".join([subject["NhomTo"], subject["MaMH"]])

                                    if key not in list_report_lms:
                                          # Lần đầu gặp key này → tạo mới entry
                                          # get_info_teacher[key][0] = mã GV, [1] = tên GV (lấy từ PHDT)
                                          list_report_lms[key] = [
                                                value_department,       # tên khoa hiển thị
                                                subject["NhomTo"],      # nhóm môn học
                                                subject["MaMH"],        # mã môn học
                                                subject["TenMH"],       # tên môn học
                                                subject["MaLop"],       # mã lớp
                                                subject["MaDP"],        # mã đơn vị
                                                subject["TenDP"],       # tên đơn vị
                                                get_info_teacher[key][0],  # mã giảng viên (từ PHDT)
                                                get_info_teacher[key][1],  # tên giảng viên (từ PHDT)
                                                from_day.strftime("%d-%m-%Y")  # ngày bắt đầu TKB
                                          ]
                                    else:
                                          # Nếu key đã tồn tại (cùng nhóm-môn nhưng khác lớp)
                                          # → gộp mã lớp vào, ngăn cách bằng dấu phẩy
                                          # Ví dụ: "TM123456,TM654321"
                                          list_report_lms[key][4] = ",".join([list_report_lms[key][4], subject["MaLop"]])

            # --- BƯỚC 5: Đối chiếu với LSA để đánh dấu môn đã soạn LMS ---
            # Duyệt qua toàn bộ kết quả, nếu key tồn tại trong danh sách LSA
            # thì append "x" (đã soạn), ngược lại append "" (chưa soạn)
            for key_sb, value_sb in list_report_lms.items():
                  if key_sb in get_group_subject_in_lsa:
                        list_report_lms[key_sb].append("x")   # đã ghi danh trên LMS
                  else:
                        list_report_lms[key_sb].append("")     # chưa ghi danh trên LMS

            # --- BƯỚC 6: Chuyển dict sang list các dict để truyền vào hàm xuất Excel ---
            # Mỗi phần tử là 1 dict với các key cố định cho dễ đọc
            list_lms_add_to_report = []
            for key, value in list_report_lms.items():
                  dict_subject = {
                        "department": value[0],    # tên khoa
                        "group": value[1],         # nhóm môn học
                        "id_subject": value[2],    # mã môn học
                        "name_subject": value[3],  # tên môn học (có thể gộp nhiều lớp)
                        "id_class": value[4],      # mã lớp
                        "id_unit": value[5],       # mã đơn vị
                        "name_unit": value[6],     # tên đơn vị
                        "id_teacher": value[7],    # mã giảng viên
                        "name_teacher": value[8],  # tên giảng viên
                        "from_day": value[9],      # ngày bắt đầu TKB (dd-mm-yyyy)
                        "has_lms": value[10]       # "x" nếu đã soạn LMS, "" nếu chưa
                  }
                  list_lms_add_to_report.append(dict_subject)

            # --- BƯỚC 7: Xuất file Excel báo cáo ---
            self.decor_report_perform_lms(list_lms_add_to_report, from_day.strftime("%d-%m-%Y"), to_day.strftime("%d-%m-%Y"))


      def set_dimension_column(self, name_sheet, title_header):
            """Tự động căn chỉnh độ rộng cột theo nội dung dài nhất trong cột đó.

            Duyệt qua tất cả ô trong từng cột, tìm giá trị dài nhất,
            sau đó set width = max_length + 10 (padding), tối đa 100 ký tự.

            Tham số:
                name_sheet  : đối tượng worksheet của openpyxl cần căn chỉnh
                title_header: list tiêu đề cột, dùng để biết có bao nhiêu cột cần xử lý
            """
            for col_num in range(1, len(title_header) + 1):
                  column_letter = get_column_letter(col_num)  # đổi số cột sang chữ cái, ví dụ 1 → "A"
                  max_length = 0

                  # Duyệt qua tất cả ô trong cột
                  for row in name_sheet[column_letter]:
                        try:
                              if len(str(row.value)) > max_length:
                                    max_length = len(str(row.value))
                        except:
                              pass

                  # Set độ rộng với padding, giới hạn tối đa 100 ký tự để tránh cột quá rộng
                  adjusted_width = min(max_length + 10, 100)
                  name_sheet.column_dimensions[column_letter].width = adjusted_width


      def decor_report_perform_lms(self, data, from_day, to_day):
            """Xuất dữ liệu báo cáo ra file Excel với định dạng đẹp.

            Tạo 2 file Excel:
              1. File báo cáo tuần (tên theo khoảng ngày):
                 - Sheet "Tổng quan": thống kê số lượng theo từng khoa
                 - Sheet "Chi tiết": danh sách chi tiết từng môn học

              2. File tích lũy "__danh_sach_khong_thuc_hien_lms.xlsx":
                 - Ghi thêm (append) các môn CHƯA soạn LMS vào cuối file
                 - File này tồn tại xuyên suốt cả học kỳ, không bị ghi đè

            Màu sắc trong sheet Chi tiết:
              - Trắng (#FFFFFF): môn đã soạn LMS
              - Vàng (#FFFF00): môn CHƯA soạn LMS → cần chú ý

            Tham số:
                data    : list các dict môn học (output từ report_perform_lms)
                from_day: ngày bắt đầu (dd-mm-yyyy) — dùng để đặt tên file
                to_day  : ngày kết thúc (dd-mm-yyyy) — dùng để đặt tên file
            """
            # Xác định đường dẫn thư mục output theo học kỳ
            base_dir = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
            folder_track_lms = os.path.join(base_dir, "data", "output", "report_perform_lms", semester)
            os.makedirs(folder_track_lms, exist_ok=True)  # tạo thư mục nếu chưa có

            # Đường dẫn file tích lũy (các môn chưa soạn LMS, ghi thêm mỗi tuần)
            flle_track_no_lms = os.path.join(folder_track_lms, "__danh_sach_khong_thuc_hien_lms.xlsx")

            # Đường dẫn file báo cáo tuần (tên chứa khoảng ngày)
            name_file = f"BC Tình hình soạn thảo LMS HK {semester} từ ngày ({from_day} đến ngày {to_day}).xlsx"
            file_report_lms = os.path.join(folder_track_lms, name_file)

            # Tạo workbook mới cho báo cáo tuần
            wb_report_lms = Workbook()
            # openpyxl tự tạo sheet tên "Sheet" mặc định → xóa đi để tự tạo lại đúng tên
            if "Sheet" in wb_report_lms.sheetnames:
                  wb_report_lms.remove(wb_report_lms["Sheet"])

            sheet_general = wb_report_lms.create_sheet("Tổng quan")

            # =========================================================
            # TÍNH THỐNG KÊ THEO KHOA cho sheet Tổng quan
            # =========================================================
            # department_stats lưu số lượng từng khoa: { tên_khoa: {count, has_x, no_x} }
            department_stats = {}
            summary_report = []
            # Hàng tổng cộng ở cuối bảng tổng quan
            footer_sheet_general = ["Tổng cộng", 0, 0, 0]

            for dic in data:
                  dept = dic['department']
                  has_lms = dic['has_lms']

                  if dept not in department_stats:
                        department_stats[dept] = {'count': 0, 'has_x': 0, 'no_x': 0}

                  department_stats[dept]['count'] += 1   # tổng số môn của khoa
                  if has_lms == 'x':
                        department_stats[dept]['has_x'] += 1   # đã soạn LMS
                  else:
                        department_stats[dept]['no_x'] += 1    # chưa soạn LMS

            # Chuyển dict thành list để ghi vào Excel
            for dept, stats in department_stats.items():
                  summary_report.append([
                        dept,
                        stats['count'],
                        stats['has_x'],
                        stats['no_x']
                  ])

            # Tính hàng tổng cộng bằng cách cộng dồn từng khoa
            for department in summary_report:
                  footer_sheet_general[1] += department[1]   # tổng số môn
                  footer_sheet_general[2] += department[2]   # tổng đã soạn
                  footer_sheet_general[3] += department[3]   # tổng chưa soạn

            # =========================================================
            # ĐỊNH NGHĨA STYLE DÙNG CHUNG CHO CẢ 2 SHEET
            # =========================================================
            header_font = Font(name="Times New Roman", size=12, bold=True)   # font tiêu đề
            footer_font = Font(name="Times New Roman", size=12, bold=True)   # font hàng tổng
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            header_fill = PatternFill(start_color="97FFFF", end_color="97FFFF", fill_type="solid")  # nền xanh nhạt cho header
            data_font = Font(name="Times New Roman", size=11)                # font dữ liệu
            data_alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
            border_style = Border(
                  left = Side(style="thin"),
                  right = Side(style="thin"),
                  top = Side(style="thin"),
                  bottom = Side(style="thin")
            )

            # =========================================================
            # GHI SHEET TỔNG QUAN
            # =========================================================
            header_sheet_general = ["Khoa", "Tổng số nhóm môn học", "Đã soạn LMS", "Chưa soạn LMS"]

            # Ghi hàng tiêu đề và áp dụng style
            sheet_general.append(header_sheet_general)
            for col_num, header in enumerate(header_sheet_general, 1):
                  cell = sheet_general.cell(row=1, column=col_num)
                  cell.font = header_font
                  cell.fill = header_fill
                  cell.alignment = header_alignment
                  cell.border = border_style

            # Ghi dữ liệu từng khoa (bắt đầu từ hàng 2)
            for row_num, row_data in enumerate(summary_report, 2):
                  for col_num, value in enumerate(row_data, 1):
                        cell = sheet_general.cell(row=row_num, column=col_num)
                        cell.value = value
                        cell.font = data_font
                        cell.alignment = data_alignment
                        cell.border = border_style

            # Ghi hàng tổng cộng ở cuối bảng tổng quan
            sheet_general.append(footer_sheet_general)
            for col_num, footer in enumerate(footer_sheet_general, 1):
                  # Hàng footer = số hàng dữ liệu + 1 hàng header + 1
                  cell = sheet_general.cell(row=len(summary_report) + 2, column=col_num)
                  cell.font = footer_font
                  cell.alignment = data_alignment
                  cell.border = border_style

            # Tự động căn chỉnh độ rộng cột cho sheet tổng quan
            self.set_dimension_column(sheet_general, header_sheet_general)

            # =========================================================
            # GHI SHEET CHI TIẾT
            # =========================================================
            sheet_detail = wb_report_lms.create_sheet("Chi tiết")

            header_sheet_detail = ["STT", "Khoa", "Nhóm môn học", "Mã môn học", "Tên môn học", "Mã lớp", "Mã đơn vị", "Tên đơn vị", "Mã giảng viên", "Tên giảng viên", "Ngày bắt đầu TKB", "Đã soạn LMS"]

            # Ghi hàng tiêu đề và áp dụng style
            sheet_detail.append(header_sheet_detail)
            for col_num, header in enumerate(header_sheet_detail, 1):
                  cell = sheet_detail.cell(row=1, column=col_num)
                  cell.font = header_font
                  cell.fill = header_fill
                  cell.alignment = header_alignment
                  cell.border = border_style

            # =========================================================
            # XỬ LÝ FILE TÍCH LŨY CÁC MÔN CHƯA SOẠN LMS
            # =========================================================
            # Nếu file chưa tồn tại (lần đầu chạy) → tạo mới với header
            # Header file tích lũy bỏ cột STT và cột "Đã soạn LMS" (vì toàn bộ đều chưa soạn)
            if not os.path.exists(flle_track_no_lms):
                  wb_new_file_track_no_lms = Workbook()
                  ws_new_file_track_no_lms = wb_new_file_track_no_lms.active
                  # header_sheet_detail[1:-1] → bỏ cột đầu (STT) và cột cuối (Đã soạn LMS)
                  ws_new_file_track_no_lms.append(header_sheet_detail[1:-1])
                  wb_new_file_track_no_lms.save(flle_track_no_lms)

            # Mở file tích lũy để ghi thêm (append) — không ghi đè toàn bộ file
            wb_file_track_no_lms = load_workbook(flle_track_no_lms)
            ws_file_track_no_lms = wb_file_track_no_lms.active
            # Tìm hàng cuối cùng hiện có, dữ liệu mới sẽ ghi từ hàng tiếp theo
            row_in_sheet_no_lms = ws_file_track_no_lms.max_row + 1

            # =========================================================
            # DUYỆT TỪNG MÔN HỌC → GHI VÀO SHEET CHI TIẾT + FILE TÍCH LŨY
            # =========================================================
            row_in_sheet_detail = 2  # bắt đầu ghi từ hàng 2 (hàng 1 là header)
            for subject in data:
                  # Màu nền mặc định: trắng (đã soạn LMS)
                  has_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

                  if subject['has_lms'] != 'x':
                        # Môn CHƯA soạn LMS → tô nền vàng để dễ nhận biết
                        has_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

                        # Ghi môn chưa soạn vào file tích lũy (ghi thêm vào cuối, không ghi đè)
                        cell_no_lms_1 = ws_file_track_no_lms.cell(row=row_in_sheet_no_lms, column=1)
                        cell_no_lms_1.value = subject['department']

                        cell_no_lms_2 = ws_file_track_no_lms.cell(row=row_in_sheet_no_lms , column=2)
                        cell_no_lms_2.value = subject['group']

                        cell_no_lms_3 = ws_file_track_no_lms.cell(row=row_in_sheet_no_lms, column=3)
                        cell_no_lms_3.value = subject['id_subject']

                        cell_no_lms_4 = ws_file_track_no_lms.cell(row=row_in_sheet_no_lms, column=4)
                        cell_no_lms_4.value = subject['name_subject']

                        cell_no_lms_5 = ws_file_track_no_lms.cell(row=row_in_sheet_no_lms, column=5)
                        cell_no_lms_5.value = subject['id_class']

                        cell_no_lms_6 = ws_file_track_no_lms.cell(row=row_in_sheet_no_lms, column=6)
                        cell_no_lms_6.value = subject['id_unit']

                        cell_no_lms_7 = ws_file_track_no_lms.cell(row=row_in_sheet_no_lms, column=7)
                        cell_no_lms_7.value = subject['name_unit']

                        cell_no_lms_8 = ws_file_track_no_lms.cell(row=row_in_sheet_no_lms, column=8)
                        cell_no_lms_8.value = subject['id_teacher']

                        cell_no_lms_9 = ws_file_track_no_lms.cell(row=row_in_sheet_no_lms, column=9)
                        cell_no_lms_9.value = subject['name_teacher']

                        cell_no_lms_10 = ws_file_track_no_lms.cell(row=row_in_sheet_no_lms, column=10)
                        cell_no_lms_10.value = subject['from_day']

                        print(f"Đã thêm môn học {subject['id_subject']} - {subject['name_subject']} - {subject['id_teacher']} - {subject['name_teacher']}  vào file theo dõi không thực hiện LMS.")
                        row_in_sheet_no_lms += 1  # chuyển sang hàng tiếp theo trong file tích lũy

                  # Ghi STT vào cột đầu tiên của sheet Chi tiết
                  cell = sheet_detail.cell(row=row_in_sheet_detail, column=1)
                  cell.value = row_in_sheet_detail - 1  # STT bắt đầu từ 1
                  cell.font = data_font
                  cell.alignment = data_alignment
                  cell.border = border_style
                  cell.fill = has_fill

                  # Ghi các trường dữ liệu của môn học vào các cột tiếp theo
                  col_in_sheet_derail = 2
                  for key, value in subject.items():
                        cell = sheet_detail.cell(row=row_in_sheet_detail, column=col_in_sheet_derail)
                        cell.value = value
                        cell.font = data_font
                        cell.alignment = data_alignment
                        cell.border = border_style
                        cell.fill = has_fill

                        col_in_sheet_derail += 1
                  row_in_sheet_detail += 1  # chuyển sang hàng tiếp theo trong sheet Chi tiết

            # Lưu file tích lũy sau khi đã ghi xong tất cả môn chưa soạn
            wb_file_track_no_lms.save(flle_track_no_lms)

            # Tự động căn chỉnh độ rộng cột cho sheet chi tiết
            self.set_dimension_column(sheet_detail, header_sheet_detail)

            # Lưu file báo cáo tuần
            wb_report_lms.save(file_report_lms)

